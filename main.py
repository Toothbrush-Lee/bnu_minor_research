#!/bin/env python

import openpyxl
import re
import numpy as np

# import pandas as pd


# major_class = {"法学": "法学",
               "心理": "社会学科类", "社会": "社会学科类", "信息": "社会学科类", "人文": "社会学科类",  # 信息管理、人文地理
               "地理": "理科类", "化学": "理科类", "环境": "理科类", "物理": "理科类", "天文": "理科类", "生物": "理科类", "资源": "理科类", "自然": "理科类",
               "俄语": "外语类", "英语": "外语类", "日语": "外语类",
               "金融": "经济金融类", "经济": "经济金融类", "会计": "经济金融类", "国际": "经济金融类",
               "特殊": "教育类", "教育": "教育类", "学前": "教育类", "思想": "教育类",
               "统计": "数统类", "数学": "数统类",
               "电子": "计算机类", "计算": "计算机类", "人工": "计算机类",
               "艺术": "艺术类", "音乐": "艺术类", "舞蹈": "艺术类", "书法": "艺术类", "戏剧": "艺术类", "美术": "艺术类",
               "数字": "新传类", "传播": "新传类",
               "人力": "管理类", "公共": "管理类", "管理": "管理类", "工商": "管理类",
               "汉语": "文科类", "历史": "文科类", "哲学": "文科类", "政治": "文科类",
               "体育": "体育类", "运动": "体育类"
               }

minor_class = {
    "哲学": "文科类", "历史": "文科类",
    '化学': '理科类', '物理': '理科类', '生物': '理科类', '地理': '理科类', '天文': '理科类', '环境': '理科类',
    '数学': '数统类', '统计': '数统类',
    '计算': '计算机类', '数据': '计算机类',
    '英语': '英语',
    '汉语': '语文',
    '教育': '教育类', '思想': '教育类', '学前': '教育类',
    '人力': '管理类', '公共': '管理类',
    '社会': '社会类',
    '传播': '传播',
    '心理': '心理',
    '法学': '法学',
    '国际': '国际经贸'
}


# 分拆两行，（分离数字）
def mov_line():
    year = input("year:")
    wb = openpyxl.load_workbook("./docs/" + year + ".xlsx")
    sheet = wb.worksheets[0]

    for cell in sheet['D']:
        y = re.sub('\D', "", cell.value)
        if y != "":
            sheet.cell(int(cell.coordinate[1:]), 5).value = y
        cell.value = re.sub('\d', "", cell.value)

    wb.save("./docs/" + year + "_new.xlsx")
    print("Numbers separated.")
    return 0


def create_dict_table(dicti, row):
    # 制作字典表
    # 新建字典
    # 实例化
    # dict_table = openpyxl.Workbook()
    # 激活 worksheet
    # dict_ws = dict_table.create_sheet("New", 0)

    dict_table = openpyxl.load_workbook('./docs/dict_table' + ".xlsx")
    dict_ws = dict_table.worksheets[0]

    count = 1
    for key in dicti.keys():
        dict_ws.cell(count, row).value = key
        dict_ws.cell(count, row + 1).value = dicti[key]
        count += 1
    dict_table.save("./docs/dict_table.xlsx")
    return 0


# 重新编码变量
def recode(filename):
    wb = openpyxl.load_workbook("./docs/" + filename + ".xlsx")
    sheet = wb.worksheets[0]

    # 制作主修专业字典：{专业:人数}
    dict_major = {}
    for cell in sheet['C']:
        # print(cell.value)
        va = cell.value[:2]
        if dict_major.__contains__(va):
            dict_major[va] += 1
        else:
            dict_major[va] = 1

    # 制作辅修专业字典：{专业:人数}
    dict_minor = {}
    for cell in sheet['F']:
        va = cell.value[:2]
        if dict_minor.__contains__(va):
            dict_minor[va] += 1
        else:
            dict_minor[va] = 1

    # 重编码主修专业字典
    dict_major1 = {}
    count = 1
    for i in dict_major.keys():
        dict_major1[i] = count
        count += 1
        # print(i)

    # 重编码辅修专业字典
    dict_minor1 = {}
    count = 1
    for i in dict_minor.keys():
        dict_minor1[i] = count
        count += 1
        # print(i)

    # print(dict_major)
    # print(dict_minor)

    # 重编码主修类别
    major_class1 = {}
    count = 1
    for i in major_class.values():
        if major_class1.__contains__(i):
            pass
        else:
            major_class1[i] = count
            count += 1
        # print(i)
    # print(major_class1)

    # 重编码辅修类别
    minor_class1 = {}
    count = 1
    for i in minor_class.values():
        if minor_class1.__contains__(i):
            pass
        else:
            minor_class1[i] = count
            count += 1
        # print(i)

    # create_dict_table(dict_major, dict_minor)
    create_dict_table(dict_major, 1)
    create_dict_table(dict_minor, 3)
    create_dict_table(major_class, 5)
    create_dict_table(minor_class, 7)

    # 制作由每个数据元组构成的列表
    # li = []

    # 实例化 recoded.xlsx
    recoded = openpyxl.Workbook()
    # 激活 worksheet
    recoded_ws = recoded.create_sheet("New", 0)
    recode_dict_a = {}
    recode_dict_i = {}
    for cell in sheet['C']:
        # print(sheet.cell(int(cell.coordinate[1:]), 6).value)
        major = cell.value
        majo = major[:2]
        recode_dict_a[major] = majo

        minor = sheet.cell(int(cell.coordinate[1:]), 6).value
        mino = minor[:2]
        recode_dict_i[minor] = mino

        majo_num = dict_major1[majo]
        mino_num = dict_minor1[mino]

        major_c = major_class[majo]
        major_cn = major_class1[major_class[majo]]
        minor_c = minor_class[mino]
        minor_cn = minor_class1[minor_class[mino]]

        # li.append( (majo_num, mino_num) )
        recoded_ws.cell(int(cell.coordinate[1:]), 1).value = majo
        recoded_ws.cell(int(cell.coordinate[1:]), 2).value = majo_num
        recoded_ws.cell(int(cell.coordinate[1:]), 3).value = mino
        recoded_ws.cell(int(cell.coordinate[1:]), 4).value = mino_num

        recoded_ws.cell(int(cell.coordinate[1:]), 5).value = int(sheet.cell(int(cell.coordinate[1:]), 4).value)  # 年级

        recoded_ws.cell(int(cell.coordinate[1:]), 6).value = major_c
        recoded_ws.cell(int(cell.coordinate[1:]), 7).value = major_cn
        recoded_ws.cell(int(cell.coordinate[1:]), 8).value = minor_c
        recoded_ws.cell(int(cell.coordinate[1:]), 9).value = minor_cn
    recoded.save("./docs/recoded.xlsx")
    create_dict_table(recode_dict_a, 9)
    create_dict_table(recode_dict_i, 11)

    # print(li)
    return 0


def get_arrey():
    recoded_book = openpyxl.load_workbook('./docs/recoded.xlsx')
    ws = recoded_book.worksheets[0]
    li = np.zeros((14, 14))     # 横排：主修；竖排：辅修
    # print(li)
    for cellG in ws["G"]:
        li[cellG.value][int(ws.cell(int(cellG.coordinate[1:]), 9).value)] += 1
    return li


def predict(X):
    X_major = X.sum(axis=1, keepdims=False)
    pre = X
    for i in range(1, 14):
        pre[i] = X[i]/X_major[i]
    # print(X)
    # print(X[1])
    # print(pre)
    # print(pre.sum(axis=1))
    return pre


def write_prediction(pre):
    # 制作预测表
    pred_table = openpyxl.load_workbook('./docs/prediction.xlsx')
    ws = pred_table.worksheets[0]
    for ro in range(1, 14):
        for co in range(1, 14):
            ws.cell(ro+1, co+1).value = pre[ro][co]
    pred_table.save("./docs/prediction.xlsx")
    print("Prediction Written.")
    return 0


def main():
    # mov_line()
    # print(len(major_class))
    recode(filename="data1")
    arr = get_arrey()
    pre = predict(arr)
    # print(pre)
    write_prediction(pre)
    print("Done!")
    return 0




main()
